from openpyxl import Workbook
import random, time
import numpy as np
import matplotlib.pyplot as plt




random.seed( time.time() )

def start():
	limits = input("Укажите максимальное и минимальное значение случайной величины\n(Например: 0, 20):\n")

	min, max = limits.split(", ")
	min = float(min)
	max = float(max)



def initTable(min=-20, max=20):

	wb = Workbook()
	ws = wb.active

	upper_limit = max
	lower_limit = min

	coords = set()

	ws['A1'] = 'x'
	ws['A2'] = 'y'

	for col in range(2, 17):
		unique_coords = False
		while not unique_coords:
			x_y = (random.uniform(lower_limit, upper_limit), random.uniform(lower_limit, upper_limit))
			if x_y not in coords:
				unique_coords = True
				coords.add(x_y)
				ws.cell(column=col, row=1, value=x_y[0])
				ws.cell(column=col, row=2, value=x_y[1])


	wb.save("Data.xlsx")

	MNK(coords)


def MNK(data):

	fig = plt.figure()

	n = len(data)

	sumx = 0
	sumx2 = 0
	sumy = 0
	sumxy = 0

	x_coords = []
	y_coords = []

	for point in data:

		plt.scatter(point[0], point[1])

		sumx += point[0]
		sumy += point[1]

		sumxy += point[0] * point[1]
		sumx2 += point[0] ** 2

		x_coords.append( point[0] )


	a = ( n * sumxy - sumx * sumy ) / ( n * sumx2 - sumx ** 2 )
	b = 1.0 * ( sumy - a * sumx ) / n


	for x in x_coords:
		y_coords.append( a * x + b )
		plt.scatter(x, a * x + b)
	
	plt.plot(x_coords, y_coords)
	plt.ylabel('Y')
	plt.xlabel('X')
	plt.show()






if __name__ == '__main__':
	start()