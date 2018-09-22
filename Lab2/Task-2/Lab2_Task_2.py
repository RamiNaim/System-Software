from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import random
import time

random.seed(time.time())

def initTable(wb, ws):
	data = set()
	for i in range(1,16):
		for j in range(1,16):
			val = random.uniform(1,30)
			ws.cell(column=j, row=i, value=val)
	val = 0
	i = 0
	while i < 10 :
		col = random.randint(1, 15)
		row = random.randint(1, 15)
		ws.cell(column=col, row=row, value=val)
		i+=1
	wb.save("Data1.xlsx")

def vinz(wb, ws):
	for i in range(1,16):
		for j in range(1,16):
			if ws['A1'].value == 0:
				ws.cell(column=1, row=1, value=15)
				continue
			if j==1 and ws[get_column_letter(j) + str(i)].value == 0:
				val = ws['O' + str(i-1)].value
				ws.cell(column=j, row=i, value=val)
				continue
			if ws[get_column_letter(j) + str(i)].value == 0: 
				val = ws[get_column_letter(j-1) + str(i)].value
				ws.cell(column=j, row=i, value=val)
	wb.save("Data1.xlsx")

def linapr(wb, ws):
	for j in range(1,16):
		sumx = 0 
		sumx2 = 0
		sumy = 0
		sumxy = 0
		for i in range(1,16):
			sumx += i-1
			sumx2 += (i-1)**2
			sumy += ws[get_column_letter(j) + str(i)].value
			sumxy += ws[get_column_letter(j) + str(i)].value*(i-1)
		a = (15*sumxy - sumx*sumy)/(15*sumx2 - sumx**2)
		b = (sumy-a*sumx)/15
		for i in range(1,16):
			if ws[get_column_letter(j) + str(i)].value == 0:
				val = a*i + b
				ws.cell(column=j, row=i, value=val)
	wb.save("Data1.xlsx")

def cor(wb, ws):
	n_str = input("Введите номер ряда, который вы хотите восстановить:\n")
	n = int(n_str)
	m_str = input("Введите номер ряда, с которым он коррелирует:\n")
	m = int(m_str)
	for j in range(1,16):
		if ws[get_column_letter(j) + str(n)].value == 0 and j==1:
			val = ws[get_column_letter(j+1) + str(n)].value/(ws[get_column_letter(j) + str(m)].value*ws[get_column_letter(j+1) + str(m)].value)
			ws.cell(column=j, row=n, value=val)			
		elif ws[get_column_letter(j) + str(n)].value == 0:
			val = ws[get_column_letter(j-1) + str(n)].value/(ws[get_column_letter(j) + str(m)].value*ws[get_column_letter(j-1) + str(m)].value)
			ws.cell(column=j, row=n, value=val)				
	wb.save("Data1.xlsx")







