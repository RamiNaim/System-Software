from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import random
from math import sqrt, log

random.seed()

mean_vals = []
var_vals = []

def initTable():
	for i in range(1,16):
		for j in range(1,16):
			val = random.uniform(1,30)
			ws.cell(column=j, row=i, value=val)
	wb.save("Data.xlsx")


def mean_and_disp():
	for i in range(1, 16):
		sq_summ = 0
		mean_vals.append( mean_row(i) )
		var_vals.append( var(i) )


def mean_row(i):
	summ = 0
	for j in range(1, 16):
			summ += ws[get_column_letter(j) + str(i)].value
	return summ / 15

def mean_list(x, y):
	summ = 0
	for i in range(15):
		summ += x[i]*y[i]

	return summ / 15


def var(i):
	sq_summ = 0
	for j in range(1, 16):
		sq_summ += ( ws[get_column_letter(j) + str(i)].value ) ** 2

	return sq_summ / 15 - mean_row(i) ** 2


def corr(i, j):
	x_var = var(i)
	y_var = var(j)

	x_centered = centered_lin(i)
	y_centered = centered_lin(j)

	x_exp_centered = centered_exp(i)
	x_exp_var = var_exp(i)

	y_exp_centered = centered_exp(j)
	y_exp_var = var_exp(j)

	r_lin = mean_list(x_centered, y_centered)
	R_lin = r_lin / ( sqrt( x_var * y_var ) )

	r_exp = mean_list(x_exp_centered, y_centered)
	R_exp = r_exp / ( sqrt( x_exp_var * y_var ) )
	"""
	r_exp_2 = mean_list(x_centered, y_exp_centered)
	R_exp_2 = r_exp_2 / ( sqrt( x_var * y_exp_var ) )

	R_exp = max(R_exp_1, R_exp_2)
	"""

	result = [R_lin, R_exp]

	return result


def var_exp(i):
	x = []
	for j in range(1, 16):
		x.append( log( ws[get_column_letter(j) + str(i)].value) )
	x_mean = mean_for_list(x)

	for j in range(15):
		x[j] = x[j] ** 2

	sq_summ = 0
	for j in range(15):
		sq_summ += x[j] ** 2

	var = sq_summ / 15 - x_mean ** 2

	return var


def centered_lin(i):
	x = []
	mean_x = mean_row(i)
	for j in range(1, 16):
		x.append( ws[get_column_letter(j) + str(i)].value - mean_x )

	return x


def centered_exp(i):
	x = []
	for j in range(1, 16):
		x.append( log( ws[get_column_letter(j) + str(i)].value) ) 

	mean_x = mean_for_list(x)
	for j in range(15):
		x[j] = x[j] - mean_x

	return x

def mean_for_list(x):
	summ = 0
	for elem in x:
		summ += elem

	return summ / 15



if __name__ == '__main__':
	wb = Workbook()
	ws = wb.active
	initTable()
	mean_and_disp()
	for i in range(15):
		print("Ряд №" + str(i+1) + " : мат.ожидание " + str(mean_vals[i]) + ", дисперсия " + str(var_vals[i]))

	corr_koeff = []

	for i in range(1, 16):
		for j in range(1, 16):
			corr_koeff.append( corr(i, j) )
			#print( str(i) + " --> " + str(j) + ": " + str( corr(i, j) ) )

	print("\n\nВведите допустимую погрешность взаимосвязи: ")
	error = float( input() )

	#print(len(corr_koeff))

	print("Линейно коррелируют ряды: ")
	for i in range(0, 15):
		for j in range(i+1, 15):
			if abs(corr_koeff[i * 15 + j][0]) > 1 - error:
				print( str(i + 1) + " --> " + str( j + 1) + ", коэф.корр = " + str(corr_koeff[i * 15 + j][0]) )


	print("\n\nЭкспоненциально коррелируют ряды: ")
	for i in range(0, 15):
		for j in range(0, 15):
			if i == j:
				continue
			if abs(corr_koeff[i * 15 + j][1]) > 1 - error:
				print( str(i + 1) + " --> " + str( j + 1) + ", коэф.корр = " + str(corr_koeff[i * 15 + j][1]) )


